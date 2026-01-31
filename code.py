import streamlit as st
import pandas as pd
import openpyxl
import io
import re

st.title("Spreadsheet Processor for Custom Columns")

uploaded_file = st.file_uploader("Upload your spreadsheet", type=["xlsx", "xls"])


# ---------- helpers ----------

def get_header_map(ws, header_row=1):
    """Map header text -> column index (1-based)."""
    header_map = {}
    for cell in ws[header_row]:
        if cell.value is not None:
            header_map[str(cell.value).strip()] = cell.col_idx
    return header_map


def is_nonempty(x):
    """True if value exists and isn't blank/NaN."""
    if x is None:
        return False
    try:
        if pd.isna(x):
            return False
    except Exception:
        pass
    return str(x).strip() != ""


def pick_source(row, prefer_col, fallback_col):
    """Prefer prefer_col if non-empty; otherwise fallback_col. Returns stripped string."""
    if prefer_col in row.index and is_nonempty(row.get(prefer_col)):
        return str(row.get(prefer_col)).strip()
    return str(row.get(fallback_col, "")).strip()


def split_color_size(spec):
    """
    Supports:
      Red/XL
      Red-XL
      Navy-Blue-XL (splits on LAST dash)
    """
    if spec is None or (isinstance(spec, float) and pd.isna(spec)):
        return "", ""

    s = str(spec).strip()
    # Normalize dash variants
    s = re.sub(r"[–—－]", "-", s)

    if "/" in s:
        c, s2 = s.split("/", 1)
        return c.strip(), s2.strip()
    if "-" in s:
        c, s2 = s.rsplit("-", 1)
        return c.strip(), s2.strip()
    return s, ""


def extract_style_code(source_id):
    """
    款号编码: ONLY ONE digit after A
    Examples:
      A2-20250703381-Navy-XL -> A2
      A8250523149R          -> A8
      A8-2025001-B          -> A8
    """
    s = "" if source_id is None else str(source_id).strip()
    s = re.sub(r"[–—－]", "-", s)
    m = re.search(r"A(\d)", s)
    return f"A{m.group(1)}" if m else "A2"


def extract_image_code(source_id):
    """
    图片编码 FINAL (dash-safe):

    1) A<digits>-<digits>...      -> return <digits> after dash
       A8-2025001-B               -> 2025001
       A2-20250703381-Navy-XL     -> 20250703381

    2) A<digits>-<letters>...     -> return A<digits>
       A820250603048-B-Black-S    -> A820250603048

    3) <digits>-<anything>...     -> return <digits>
       20250205010-White-L        -> 20250205010

    4) Otherwise                  -> return full string
       A8250523149R               -> A8250523149R
    """
    if source_id is None:
        return ""

    s = str(source_id).strip()
    # Normalize all dash types to normal hyphen
    s = re.sub(r"[–—－]", "-", s)

    # Case 1: A...-<digits>
    m1 = re.match(r"^A\d+-(\d+)", s)
    if m1:
        return m1.group(1)

    # Case 2: A<digits>-<anything>
    m2 = re.match(r"^(A\d+)-", s)
    if m2:
        return m2.group(1)

    # Case 3: <digits>-<anything>
    m3 = re.match(r"^(\d+)[-_]", s)
    if m3:
        return m3.group(1)

    return s


# ---------- main ----------

if uploaded_file is not None:
    try:
        # Load workbook for writing (preserves formatting)
        workbook = openpyxl.load_workbook(uploaded_file)
        sheet = workbook.active

        # Reset pointer for pandas
        uploaded_file.seek(0)
        df = pd.read_excel(uploaded_file, engine="openpyxl")

        st.write("### Original Spreadsheet")
        st.dataframe(df)

        # Required source columns
        if "规格属性" not in df.columns or "SKCID" not in df.columns:
            st.error("Uploaded file must contain '规格属性' and 'SKCID' columns.")
            st.stop()

        # Map headers in the Excel sheet (prevents column shifting)
        header_map = get_header_map(sheet, header_row=1)

        target_headers = ["*款号编码", "*颜色编码", "*尺寸编码", "*图片编码", "*工艺类型"]
        missing = [h for h in target_headers if h not in header_map]
        if missing:
            st.error(f"Missing required target columns: {missing}")
            st.stop()

        col_style = header_map["*款号编码"]
        col_color = header_map["*颜色编码"]
        col_size  = header_map["*尺寸编码"]
        col_img   = header_map["*图片编码"]
        col_proc  = header_map["*工艺类型"]

        for index, row in df.iterrows():
            # Prefer 商家编码, else SKCID
            source_id = pick_source(row, "商家编码", "SKCID")

            款号编码 = extract_style_code(source_id)
            图片编码 = extract_image_code(source_id)
            颜色编码, 尺寸编码 = split_color_size(row.get("规格属性"))

            excel_row = index + 2  # header row = 1

            sheet.cell(excel_row, col_style, 款号编码)
            sheet.cell(excel_row, col_color, 颜色编码)
            sheet.cell(excel_row, col_size,  尺寸编码)
            sheet.cell(excel_row, col_img,   图片编码)
            sheet.cell(excel_row, col_proc,  "白墨烫画")

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        st.success("Processing complete — 图片编码 now also handles leading-digit IDs.")

        st.download_button(
            "Download Processed Spreadsheet",
            buffer,
            "processed_spreadsheet.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"An unexpected error occurred: {e}")
